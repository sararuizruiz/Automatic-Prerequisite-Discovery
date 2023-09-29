# -*- coding: utf-8 -*-
"""
Created on Mon Apr 20 18:13:39 2020

@author: Sara
"""

import xlrd
import xlsxwriter
from openpyxl import load_workbook

umbral = 0.53  #Preguntar a Bea, qué umbral pongo para considerar que es prerequisito
#He tomado que si P(noB/noA)>umbral entonces A es prerequisito de B.

libro = load_workbook('discretizado1005.xlsx')
hoja = libro['Sheet1']

"""
libro = xlsxwriter.Workbook('sinbucles1005.xlsx')
hoja = libro.add_worksheet()

for i in range(hojacon.nrows):
    for j in range(hojacon.ncols):
        hoja.write(i,j,hojacon.cell_value(i,j))
"""
numalumnos = hoja.max_row-1       
haybucle = True
while (haybucle):
    
    numtemas = hoja.max_column
    
    combinaciones = [] # Contiene [[0,1],[1,0],[0,2],...] Así todas las pos combinaciones de temas que tenemos que tener en cuenta. 0 representa el tema representado en la columna 0 y así.
    for i in range(numtemas):
        for j in range(i+1,numtemas):
            combinaciones.append([i,j])
            combinaciones.append([j,i])
    
    prerequisitos = [] #Prerequisitos va a ser una lista que contiene:
                   # Para combinacion[i]=[A,B], prerequisitos[i] contiene [[A,B],[0]] si A no es un prerequisito de B
                   # y contiene [[A,B],[1,0.8,0.2]] si A es prerequisito de B y además si P(B/A)=0.8, P(B/noA)=0.2
    for [t1,t2] in combinaciones:
        numdeunos = 0
        num1 = 0.
        numdeceros = 0
        num0 = 0.
        for fila in range(numalumnos):
            if hoja.cell(row=fila+2,column=t1+1).value==1:
                numdeunos+=1
                if hoja.cell(row=fila+2,column=t2+1).value==1:
                    num1+=1
            else:
                numdeceros+=1
                if hoja.cell(row=fila+2,column=t2+1).value==0:
                    num0+=1
        if (num0/numdeceros)>umbral:
            prerequisitos.append([[t1,t2],[1]])
        else:
            prerequisitos.append([[t1,t2],[0]])
    
   # print(prerequisitos)
       
    ## Ahora vamos a evitar que A sea prerequisito de B a la vez que B de A, si pasa esto los juntamos como un único tema:
    
    for i in range(int(len(prerequisitos)/2)):
        [[t1,t2],[result1]] = prerequisitos[2*i]
        [[t2,t1],[result2]] = prerequisitos[2*i+1]
        if (result1==1 and result2==1):
            acum1 = hoja.cell(row=1,column=t1+1).value  ##Aquí creo que puedo quitar todos los +1 si la variable la empiezo a contar en 1 en vez de en 0
            acum2 = hoja.cell(row=1,column=t2+1).value
            hoja.cell(row=1,column=t1+1).value = acum1+','+acum2 #Escribo en el titulo del tema que junto a t1 y t2
            for fila in range(numalumnos):
                acum1 = hoja.cell(row=fila+2,column=t1+1).value
                acum2 = hoja.cell(row=fila+2,column=t2+1).value
                hoja.cell(row=fila+2,column=t1+1).value = acum1*acum2 ##CRITERI0: Multiplicación de los datos de los temas t1 y t2
            hoja.delete_cols(t2+1)
            break     #Aquí ahora dentro del if para el bucle for, pues ya he encontrado un bucle y lo he arreglado. Y con el nuevo excel, necesito repetir el algoritmo.
    else:  #Si consigo terminar todo el bucle for sin que haya bucles de dos, sin entrar en el if y sufrir un break, entonces ya no hay bucles
        haybucle = False

libro.save('discretizado1005.xlsx')
         
## Faltaría comprobar que no haya bucles de más de dos temas involucrados. Ej: el 1 prereq del 2, el 2 del 3 y el 3 del 1.