# -*- coding: utf-8 -*-
"""
Created on Sun May 24 13:36:35 2020

@author: Sara
"""


from csacoprerequisitossub import *
import xlrd
import xlsxwriter
from openpyxl import load_workbook


probfinal = []
padresfinal = []

temaspadre = [18768,24425,24426,24427,20947]
for indice in range (len(temaspadre)):
    
    libro = load_workbook('discretizado'+str(temaspadre[indice])+'.xlsx')
    hoja = libro['Sheet1']
        
    numtemas = hoja.max_column-1
    numalumnos = hoja.max_row-1
    padres = []
    prerequisitos = prere[indice]
    
    for i in range(numtemas):
        temp = []
        for [[t1,t2],[j]] in prerequisitos:
            if (t2==i and j==1):
                temp.append(t1)
        padres.append([[i],temp])
    padresfinal.append(padres)
        
    
    posibilidades = []  
    for [[t1],padrest1] in padres:
        pos = [[]]
        for padre in padrest1:
            copia = pos.copy()
            for i in range(len(copia)): 
                ex = pos.pop(0)
                extra = ex.copy()
                ex.append(0)
                extra.append(1)
                pos.append(ex)
                pos.append(extra)       
        posibilidades.append([[t1],pos])
    
    probabilidades = []
    for i in range(numtemas):
        [[t1],padrest1] = padres[i]
        [[t1],post1] = posibilidades[i]
        prob = []
        for pos in post1:
            numpos = 0.
            numtotal = 0.
            for fila in range(numalumnos):
                acum = True
                for j in range(len(padrest1)):
                    if hoja.cell(row=fila+2,column=padrest1[j]+2).value!=pos[j]:
                        acum = False
                if acum and type(hoja.cell(row=fila+2,column=t1+2).value)==int:
                    numtotal += 1
                    if hoja.cell(row=fila+2,column=t1+2).value==1:
                        numpos +=1
            if numtotal==0:
                prob.append(0.5)
            else:
                prob.append(numpos/numtotal)
        probabilidades.append([[t1],prob])
    print(probabilidades)

    probfinal.append(probabilidades)
    
                