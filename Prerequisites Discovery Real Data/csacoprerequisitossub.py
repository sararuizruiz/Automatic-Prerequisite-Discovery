# -*- coding: utf-8 -*-
"""
Created on Tue May 19 21:48:30 2020

@author: Sara
"""


import xlrd
import xlsxwriter
from openpyxl import load_workbook

umbral = 0.75

prere = []

temaspadre = [18768,24425,24426,24427,20947]
for padres in temaspadre:

    libro = load_workbook('discretizado'+str(padres)+'.xlsx')
    hoja = libro['Sheet1']
    
    numalumnos = hoja.max_row-1      
    haybucle = True
    while (haybucle):
        
        numtemas = hoja.max_column-1
        
        combinaciones = []
        for i in range(numtemas):
            for j in range(i+1,numtemas):
                combinaciones.append([i,j])
                combinaciones.append([j,i])
        
        prerequisitos = []  
        for [t1,t2] in combinaciones:
            numdeceros = 0
            num0 = 0.
            for fila in range(numalumnos):
                if type(hoja.cell(row=fila+2,column=t1+2).value)==int and type(hoja.cell(row=fila+2,column=t2+2).value)==int :
                    if hoja.cell(row=fila+2,column=t1+2).value==0:
                        numdeceros+=1
                        if hoja.cell(row=fila+2,column=t2+2).value==0:
                            num0+=1
    
            if numdeceros != 0:
                prob = num0/numdeceros
            else:
                prob = 0.5
            if prob>umbral:
                prerequisitos.append([[t1,t2],[1]])
            else:
                prerequisitos.append([[t1,t2],[0]])
               
        for i in range(int(len(prerequisitos)/2)):
            [[t1,t2],[result1]] = prerequisitos[2*i]
            [[t2,t1],[result2]] = prerequisitos[2*i+1]
            if (result1==1 and result2==1):
                acum1 = hoja.cell(row=1,column=t1+2).value
                acum2 = hoja.cell(row=1,column=t2+2).value
                hoja.cell(row=1,column=t1+2).value = str(acum1)+'_'+str(acum2) 
                for fila in range(numalumnos):
                    acum1 = hoja.cell(row=fila+2,column=t1+2).value
                    acum2 = hoja.cell(row=fila+2,column=t2+2).value
                    if type(acum1)==int and type(acum2)==int:
                        hoja.cell(row=fila+2,column=t1+2).value = acum1*acum2
                    elif type(acum1)==str and type(acum2)==int:
                        hoja.cell(row=fila+2,column=t1+2).value = acum2
                    elif type(acum1)==int and type(acum2)==str:
                        hoja.cell(row=fila+2,column=t1+2).value = acum1
                    elif type(acum1)==str and type(acum2)==str:
                        hoja.cell(row=fila+2,column=t1+2).value = acum1
                        
                hoja.delete_cols(t2+2)
                break
        else:  
            haybucle = False
        
        prere.append(prerequisitos)
    libro.save('discretizado'+str(padres)+'.xlsx')
             
  