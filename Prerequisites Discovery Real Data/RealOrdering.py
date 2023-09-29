# -*- coding: utf-8 -*-
"""
Created on Sun May  3 12:59:49 2020

@author: Sara
"""

## LA PRIMERA FILA ES EL IDENTIFICADOR DE CADA TEMA, NO LOS DATOS DE UN ALUMNO.
import xlrd
import xlsxwriter
import ast
from openpyxl import load_workbook

# A tener en cuenta: cell_value empieza la cuenta en 0 y cell().value empieza la cuenta en 1

librodatos = xlrd.open_workbook('datosreales.xls')
hojadatos = librodatos.sheet_by_index(0)

librodiscretizado = xlsxwriter.Workbook('ordenadoreal.xlsx')
hojadiscretizada = librodiscretizado.add_worksheet()
hojaindices = librodiscretizado.add_worksheet()

librodiscretizado.close()

librodiscretizado = load_workbook('ordenadoreal.xlsx')
hojadiscretizada = librodiscretizado['Sheet1']
hojaindices = librodiscretizado['Sheet2']

filas = hojadatos.nrows-1
#(Número de filas con información, quito la 1 que son los títulos)

numtemas=5 
hojadiscretizada.cell(1,2).value='18768'
hojadiscretizada.cell(1,3).value='24425'
hojadiscretizada.cell(1,4).value='24426'
hojadiscretizada.cell(1,5).value='24427'
hojadiscretizada.cell(1,6).value='20947'



for fila in range(1,filas+1):
    alumno = hojadatos.cell_value(fila,1)
    tema = hojadatos.cell_value(fila,7)
    calif = hojadatos.cell_value(fila,11)
    numalumnos = hojadiscretizada.max_row-1
    for i in range(1,numalumnos+1):
        if hojadiscretizada.cell(i+1,1).value==alumno:
            for j in range(1,numtemas+1):
                if hojadiscretizada.cell(1,j+1).value==str(int(tema)):
                    puntos = hojadiscretizada.cell(i+1,j+1).value
                    total = hojaindices.cell(i+1,j+1).value
                    if puntos is not None:
                        hojadiscretizada.cell(i+1,j+1).value = puntos+calif
                        hojaindices.cell(i+1,j+1).value = total+1
                    else:
                        hojadiscretizada.cell(i+1,j+1).value = calif
                        hojaindices.cell(i+1,j+1).value = 1.0
                    break
            break
    else:
        hojadiscretizada.cell(numalumnos+2,1).value = alumno
        for j in range(1,numtemas+1):
            if hojadiscretizada.cell(1,j+1).value==str(int(tema)):
                hojadiscretizada.cell(numalumnos+2,j+1).value = float(calif)
                hojaindices.cell(numalumnos+2,j+1).value = 1.0
                break
            

## Ahora ya tenemos cada una de las hojas creadas (sólo falta dividir para obtener la media aritmética):
                # Hoja1: la suma de todas las calificaciones
                # Hoja2: Número total de calificaciones

nfilas = hojadiscretizada.max_row-1 
ncolumn = hojadiscretizada.max_column-1
for i in range(2,nfilas+2):
    for j in range(2,ncolumn+2):
        sumanotas = hojadiscretizada.cell(i,j).value
        sumatotal = hojaindices.cell(i,j).value
        hojadiscretizada.cell(i,j).value = sumanotas/sumatotal


librodiscretizado.save('ordenadoreal.xlsx')
