# -*- coding: utf-8 -*-
"""
Created on Sun May  3 12:59:49 2020

@author: Sara
"""

import xlrd
import xlsxwriter
import ast
from openpyxl import load_workbook

librodatos = xlrd.open_workbook('datosreales.xls')
hojadatos = librodatos.sheet_by_index(0)
filas = hojadatos.nrows-1

temaspadre = [18768,24425,24426,24427,20947]
for padres in temaspadre:

    librodiscretizado = xlsxwriter.Workbook('ordenado'+str(padres)+'.xlsx')
    hojadiscretizada = librodiscretizado.add_worksheet()
    hojaindices = librodiscretizado.add_worksheet()
    
    librodiscretizado.close()
    
    librodiscretizado = load_workbook('ordenado'+str(padres)+'.xlsx')
    hojadiscretizada = librodiscretizado['Sheet1']
    hojaindices = librodiscretizado['Sheet2']
    
    for fila in range(1,filas+1):
        alumno = hojadatos.cell_value(fila,1)
        temapadre = hojadatos.cell_value(fila,7)
        tema = hojadatos.cell_value(fila,5)
        calif = hojadatos.cell_value(fila,11)
        if temapadre == float(padres):
            numalumnos = hojadiscretizada.max_row-1
            numtemas = hojadiscretizada.max_column-1
            for i in range(1,numalumnos+1):
                if hojadiscretizada.cell(i+1,1).value==alumno:
                    for j in range(1,numtemas+1):
                        if hojadiscretizada.cell(1,j+1).value==tema:
                            puntos = hojadiscretizada.cell(i+1,j+1).value
                            total = hojaindices.cell(i+1,j+1).value
                            if puntos is not None:
                                hojadiscretizada.cell(i+1,j+1).value = puntos+calif
                                hojaindices.cell(i+1,j+1).value = total+1
                            else:
                                hojadiscretizada.cell(i+1,j+1).value = calif
                                hojaindices.cell(i+1,j+1).value = 1.0
                            break  
                    else:
                        hojadiscretizada.cell(1,numtemas+2).value = tema
                        hojadiscretizada.cell(i+1,numtemas+2).value = calif
                        hojaindices.cell(i+1,numtemas+2).value = 1.0
                    break
            else:
                hojadiscretizada.cell(numalumnos+2,1).value = alumno
                for j in range(1,numtemas+1):
                    if hojadiscretizada.cell(1,j+1).value==tema:
                        hojadiscretizada.cell(numalumnos+2,j+1).value = float(calif)
                        hojaindices.cell(numalumnos+2,j+1).value = 1.0
                        break
                else:
                    hojadiscretizada.cell(1,numtemas+2).value = tema
                    hojadiscretizada.cell(numalumnos+2,numtemas+2).value = calif
                    hojaindices.cell(numalumnos+2,numtemas+2).value = 1.0
                    
    nfilas = hojadiscretizada.max_row-1 
    ncolumn = hojadiscretizada.max_column-1
    for i in range(2,nfilas+2):
        for j in range(2,ncolumn+2):
            sumanotas = hojadiscretizada.cell(i,j).value
            sumatotal = hojaindices.cell(i,j).value
            if sumanotas is not None:
                hojadiscretizada.cell(i,j).value = sumanotas/sumatotal
    
    librodiscretizado.remove_sheet(hojaindices)
    librodiscretizado.save('ordenado'+str(padres)+'.xlsx')

